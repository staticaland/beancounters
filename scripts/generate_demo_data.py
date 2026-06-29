"""Generate the canonical Beancounters demo statement files."""

from __future__ import annotations

import argparse
import calendar
import csv
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from lxml import etree as ET
from openpyxl import Workbook, load_workbook


YEAR = 2025
OVERLAP_START = date(YEAR, 2, 15)
OVERLAP_END = date(YEAR, 4, 15)
CHECKING = "12345678901"
SAVINGS = "11112222333"
SALARY = "56712345678"
DNB = "22223333444"
AMEX = "33334444555"
GENERATED_LEDGER_DIR = Path("generated")


@dataclass(frozen=True)
class BankTransaction:
    date: date
    description: str
    amount: Decimal
    to_account: str
    from_account: str


@dataclass(frozen=True)
class CardTransaction:
    date: date
    description: str
    amount: Decimal
    memo: str
    fitid: str


@dataclass(frozen=True)
class MortgageTransaction:
    date: date
    description: str
    principal: Decimal
    interest: Decimal

    @property
    def payment(self) -> Decimal:
        return self.principal + self.interest


def money(value: str | int) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def d(year: int, month: int, day: int) -> date:
    return date(year, month, min(day, calendar.monthrange(year, month)[1]))


def month_end(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


def in_range(tx_date: date, start: date, end: date) -> bool:
    return start <= tx_date <= end


def signed_total(transactions: list[CardTransaction]) -> Decimal:
    return sum((tx.amount for tx in transactions), Decimal("0.00"))


def mortgage_transactions(year: int) -> list[MortgageTransaction]:
    rows = [
        MortgageTransaction(d(year, month, 2), f"Mortgage Payment {month_end(year, month).strftime('%B')} {year}", money(5600 + (month - 1) * 35), money(6900 - (month - 1) * 35))
        for month in range(1, 13)
    ]
    rows.append(MortgageTransaction(d(year, 6, 15), "Extra Mortgage Payment - Vacation bonus", money(50000), money(0)))
    rows.append(MortgageTransaction(d(year, 12, 20), "Extra Mortgage Payment - Year-end savings", money(100000), money(0)))
    return sorted(rows, key=lambda tx: tx.date)


def bank_transactions_for_month(year: int, month: int, dnb_payment: Decimal, amex_payment: Decimal) -> list[BankTransaction]:
    salary = money(43500 + (month % 4) * 375)
    grocery_offset = Decimal(month * 17)
    rent = money(17800 if month < 8 else 18400)
    savings = money(6500 if month not in {7, 12} else 4000)

    rows = [
        BankTransaction(d(year, month, 1), f"HUSLEIE {month_end(year, month).strftime('%B').upper()}", -rent, "65432198765", CHECKING),
        BankTransaction(d(year, month, 3), "GET/TELIA", money("-749.00"), "34512389012", CHECKING),
        BankTransaction(d(year, month, 5), "KIWI MAJORSTUEN", -(money("612.40") + grocery_offset), "98765432109", CHECKING),
        BankTransaction(d(year, month, 7), "RUTER MANEDSKORT", money("-897.00"), "34567891234", CHECKING),
        BankTransaction(d(year, month, 10), "COOP EXTRA GRONLAND", -(money("884.20") + grocery_offset / 2), "78901234567", CHECKING),
        BankTransaction(d(year, month, 12), "SPOTIFY", money("-129.00"), "11223344556", CHECKING),
        BankTransaction(d(year, month, 14), "Lonn KOMPLETT AS", salary, CHECKING, SALARY),
        BankTransaction(d(year, month, 16), "Kafe Oslo", money("-96.00"), "44556677889", CHECKING),
        BankTransaction(d(year, month, 18), "MENY BOGSTADVEIEN", -(money("542.35") + grocery_offset), "23451298765", CHECKING),
        BankTransaction(d(year, month, 23), "Overforing til Sparekonto", -savings, SAVINGS, CHECKING),
        BankTransaction(d(year, month, 25), "NETFLIX", money("-179.00"), "22334455667", CHECKING),
        BankTransaction(d(year, month, 27), "REMA 1000 TORSHOV", -(money("731.80") + grocery_offset / 3), "33445566778", CHECKING),
        BankTransaction(d(year, month, 28), "FINN.NO FAKTURA", money("-149.00"), "87654321987", CHECKING),
    ]
    if dnb_payment:
        rows.append(BankTransaction(d(year, month, 20), "DNB MASTERCARD FAKTURA", -dnb_payment, DNB, CHECKING))
    if amex_payment:
        rows.append(BankTransaction(d(year, month, 21), "AMEX AUTOGIRO", -amex_payment, AMEX, CHECKING))
    if month in {2, 5, 9, 12}:
        rows.append(BankTransaction(d(year, month, 9), "VINMONOPOLET OSLO S", money("-689.00"), "89123456789", CHECKING))
    if month in {3, 6, 10}:
        rows.append(BankTransaction(d(year, month, 29), "STATOIL FURUSET", money("-812.45"), "23489156123", CHECKING))
    if month == 6:
        rows.append(BankTransaction(d(year, month, 24), "SKATTEETATEN", money("3150.00"), CHECKING, "99988877766"))
    if month in {4, 11}:
        rows.append(BankTransaction(d(year, month, 26), "POWER STORO", money("-3290.00"), "56712389012", CHECKING))
    return sorted(rows, key=lambda tx: tx.date, reverse=True)


def dnb_transactions_for_month(year: int, month: int, payment: Decimal) -> list[CardTransaction]:
    base = [
        (2, "XXL SPORT ALNA", "-1499.00", "Sports equipment"),
        (4, "KIWI MAJORSTUEN", "-486.70", "Groceries"),
        (6, "GITHUB.COM", "-129.00", "Developer subscription"),
        (8, "STARBUCKS KARL JOHAN", "-94.00", "Coffee"),
        (11, "REMA 1000 TORSHOV", "-738.20", "Groceries"),
        (13, "NETFLIX.COM", "-179.00", "Streaming service"),
        (17, "POWER STORO", "-2199.00", "Electronics"),
        (19, "VINMONOPOLET OSLO S", "-459.00", "Wine purchase"),
        (24, "MENY BOGSTADVEIEN", "-681.55", "Groceries"),
    ]
    rows = [
        CardTransaction(d(year, month, day), desc, money(amount) - Decimal(month * 3 if "KIWI" in desc or "MENY" in desc else 0), memo, f"DNB-{year}{month:02d}-{i:03d}")
        for i, (day, desc, amount, memo) in enumerate(base, 1)
    ]
    if month in {3, 7, 11}:
        rows.append(CardTransaction(d(year, month, 22), "SAS EUROBONUS", money("-3290.00"), "Flights", f"DNB-{year}{month:02d}-101"))
    if month in {5, 10}:
        rows.append(CardTransaction(d(year, month, 26), "XXL SPORT ALNA REFUND", money("399.00"), "Returned item", f"DNB-{year}{month:02d}-102"))
    if payment:
        rows.append(CardTransaction(d(year, month, 20), "Innbetaling", payment, "Payment from SpareBank 1", f"DNB-{year}{month:02d}-PAY"))
    return sorted(rows, key=lambda tx: tx.date, reverse=True)


def amex_transactions_for_month(year: int, month: int, payment: Decimal) -> list[CardTransaction]:
    base = [
        (3, "ELKJOP STORO", "-1299.00", "Electronics purchase"),
        (5, "SPOTIFY AB", "-129.00", "Monthly subscription"),
        (9, "H&M OSLO CITY", "-849.00", "Clothing"),
        (12, "KIWI 587 MAJORSTUEN", "-332.50", "Groceries"),
        (15, "STARBUCKS AKER BRYGGE", "-92.00", "Coffee"),
        (18, "REMA 1000 TORSHOV", "-611.40", "Groceries"),
        (23, "SAS EUROBONUS", "-2490.00", "Travel"),
    ]
    rows = [
        CardTransaction(d(year, month, day), desc, money(amount) - Decimal(month * 2 if "KIWI" in desc or "REMA" in desc else 0), memo, f"AMEX-{year}{month:02d}-{i:03d}")
        for i, (day, desc, amount, memo) in enumerate(base, 1)
    ]
    if month in {2, 6, 10}:
        rows.append(CardTransaction(d(year, month, 20), "VINMONOPOLET AKER BRYGGE", money("-529.00"), "Wine purchase", f"AMEX-{year}{month:02d}-101"))
    if month in {4, 9}:
        rows.append(CardTransaction(d(year, month, 26), "ELKJOP STORO REFUND", money("499.00"), "Returned accessory", f"AMEX-{year}{month:02d}-102"))
    if payment:
        rows.append(CardTransaction(d(year, month, 21), "AUTOGIROBETALING", payment, "Payment from SpareBank 1", f"AMEX-{year}{month:02d}-PAY"))
    return sorted(rows, key=lambda tx: tx.date, reverse=True)


def build_scenario(year: int) -> tuple[list[BankTransaction], list[CardTransaction], list[CardTransaction]]:
    dnb_by_month: dict[int, list[CardTransaction]] = {}
    amex_by_month: dict[int, list[CardTransaction]] = {}
    bank_rows: list[BankTransaction] = []

    previous_dnb = money("0")
    previous_amex = money("0")
    for month in range(1, 13):
        dnb_payment = -previous_dnb if previous_dnb < 0 else money("0")
        amex_payment = -previous_amex if previous_amex < 0 else money("0")
        dnb_rows = dnb_transactions_for_month(year, month, dnb_payment)
        amex_rows = amex_transactions_for_month(year, month, amex_payment)
        bank_rows.extend(bank_transactions_for_month(year, month, dnb_payment, amex_payment))
        dnb_by_month[month] = dnb_rows
        amex_by_month[month] = amex_rows
        previous_dnb = signed_total([tx for tx in dnb_rows if "PAY" not in tx.fitid])
        previous_amex = signed_total([tx for tx in amex_rows if "PAY" not in tx.fitid])

    return bank_rows, [tx for rows in dnb_by_month.values() for tx in rows], [tx for rows in amex_by_month.values() for tx in rows]


def format_nok(amount: Decimal) -> str:
    return f"{amount:.2f}".replace(".", ",")


def write_sparebank1(path: Path, rows: list[BankTransaction]) -> None:
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file, delimiter=";", quoting=csv.QUOTE_ALL)
        file.write("Dato;Beskrivelse;Rentedato;Inn;Ut;Til konto;Fra konto;\n")
        for tx in rows:
            writer.writerow(
                [
                    tx.date.strftime("%d.%m.%Y"),
                    tx.description,
                    "",
                    format_nok(tx.amount) if tx.amount > 0 else "",
                    format_nok(tx.amount) if tx.amount < 0 else "",
                    tx.to_account,
                    tx.from_account,
                    "",
                ]
            )


def write_dnb(path: Path, rows: list[CardTransaction]) -> None:
    wb = Workbook()
    generated_at = datetime(2025, 1, 1, tzinfo=timezone.utc)
    wb.properties.creator = "beancounters demo data generator"
    wb.properties.lastModifiedBy = "beancounters demo data generator"
    wb.properties.created = generated_at
    wb.properties.modified = generated_at
    ws = wb.active
    ws.title = "DNB Mastercard Demo"
    ws.append(["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"])
    for tx in rows:
        ws.append([tx.date, tx.description, None, None, float(tx.amount) if tx.amount > 0 else None, float(-tx.amount) if tx.amount < 0 else None])
    wb.save(path)
    normalize_xlsx(path, generated_at)


def normalize_xlsx(path: Path, generated_at: datetime) -> None:
    fixed_time = (generated_at.year, generated_at.month, generated_at.day, 0, 0, 0)
    fixed_timestamp = generated_at.strftime("%Y-%m-%dT%H:%M:%SZ")
    entries: list[tuple[ZipInfo, bytes]] = []

    with ZipFile(path, "r") as source:
        for source_info in source.infolist():
            data = source.read(source_info.filename)
            if source_info.filename == "docProps/core.xml":
                root = ET.fromstring(data)
                modified = root.find("{http://purl.org/dc/terms/}modified")
                if modified is not None:
                    modified.text = fixed_timestamp
                data = ET.tostring(root, encoding="utf-8")

            info = ZipInfo(source_info.filename, fixed_time)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = source_info.external_attr
            entries.append((info, data))

    tmp_path = path.with_suffix(".tmp")
    with ZipFile(tmp_path, "w") as target:
        for info, data in entries:
            target.writestr(info, data)
    tmp_path.replace(path)


def qbo_timestamp(value: date) -> str:
    return value.strftime("%Y%m%d000000")


def write_amex(path: Path, rows: list[CardTransaction], start: date, end: date) -> None:
    ofx = ET.Element("OFX")
    signon = ET.SubElement(ET.SubElement(ofx, "SIGNONMSGSRSV1"), "SONRS")
    status = ET.SubElement(signon, "STATUS")
    ET.SubElement(status, "CODE").text = "0"
    ET.SubElement(status, "SEVERITY").text = "INFO"
    ET.SubElement(signon, "DTSERVER").text = qbo_timestamp(end)
    ET.SubElement(signon, "LANGUAGE").text = "ENG"
    ET.SubElement(ET.SubElement(signon, "FI"), "ORG").text = "AMEX"
    ccstmt = ET.SubElement(ET.SubElement(ET.SubElement(ET.SubElement(ofx, "CREDITCARDMSGSRSV1"), "CCSTMTTRNRS"), "CCSTMTRS"), "BANKTRANLIST")
    ET.SubElement(ccstmt, "DTSTART").text = qbo_timestamp(start)
    ET.SubElement(ccstmt, "DTEND").text = qbo_timestamp(end)
    for tx in rows:
        trn = ET.SubElement(ccstmt, "STMTTRN")
        ET.SubElement(trn, "TRNTYPE").text = "CREDIT" if tx.amount > 0 else "DEBIT"
        ET.SubElement(trn, "DTPOSTED").text = qbo_timestamp(tx.date)
        ET.SubElement(trn, "TRNAMT").text = f"{tx.amount:.2f}"
        ET.SubElement(trn, "FITID").text = tx.fitid
        ET.SubElement(trn, "NAME").text = tx.description
        ET.SubElement(trn, "MEMO").text = tx.memo

    xml = ET.tostring(ofx, encoding="unicode", pretty_print=True)
    path.write_text(
        '<?xml version="1.0" standalone="no"?>\n'
        '<?OFX OFXHEADER="200" VERSION="202" SECURITY="NONE" OLDFILEUID="NONE" NEWFILEUID="NONE"?>\n'
        f"{xml}\n",
        encoding="utf-8",
    )


def write_mortgage_ledger(path: Path, year: int) -> None:
    rows = mortgage_transactions(year)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        ";; -*- mode: beancount -*-",
        f"; Generated by scripts/generate_demo_data.py for {year} mortgage insights.",
        "",
        f"{year - 1}-12-31 * \"Mortgage - Opening loan balance\"",
        "  Liabilities:Loan:Mortgage       -1850000.00 NOK",
        "  Equity:Opening-Balances",
        "",
        f"{year}-01-01 balance Liabilities:Loan:Mortgage -1850000.00 NOK",
        "",
    ]
    for tx in rows:
        lines.extend(
            [
                f"{tx.date.isoformat()} * \"{tx.description}\"",
                f"  Assets:Bank:SpareBank1:Checking  -{tx.payment:.2f} NOK",
                f"  Liabilities:Loan:Mortgage          {tx.principal:.2f} NOK",
            ]
        )
        if tx.interest:
            lines.append(f"  Expenses:Interest:Mortgage         {tx.interest:.2f} NOK")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


def ensure_clean_provider_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    for child in path.iterdir():
        if child.is_file():
            child.unlink()


def emit(output_root: Path, year: int) -> None:
    bank_rows, dnb_rows, amex_rows = build_scenario(year)
    providers = {
        "sparebank1": (".csv", bank_rows),
        "dnb": (".xlsx", dnb_rows),
        "amex": (".qbo", amex_rows),
    }
    for provider in providers:
        ensure_clean_provider_dir(output_root / provider)

    for month in range(1, 13):
        start = date(year, month, 1)
        end = month_end(year, month)
        write_sparebank1(output_root / "sparebank1" / f"{year}-{month:02d}.csv", [tx for tx in bank_rows if in_range(tx.date, start, end)])
        write_dnb(output_root / "dnb" / f"{year}-{month:02d}.xlsx", [tx for tx in dnb_rows if in_range(tx.date, start, end)])
        write_amex(output_root / "amex" / f"{year}-{month:02d}.qbo", [tx for tx in amex_rows if in_range(tx.date, start, end)], start, end)

    overlap_name = f"{OVERLAP_START.isoformat()}_to_{OVERLAP_END.isoformat()}"
    write_sparebank1(output_root / "sparebank1" / f"{overlap_name}.csv", [tx for tx in bank_rows if in_range(tx.date, OVERLAP_START, OVERLAP_END)])
    write_dnb(output_root / "dnb" / f"{overlap_name}.xlsx", [tx for tx in dnb_rows if in_range(tx.date, OVERLAP_START, OVERLAP_END)])
    write_amex(output_root / "amex" / f"{overlap_name}.qbo", [tx for tx in amex_rows if in_range(tx.date, OVERLAP_START, OVERLAP_END)], OVERLAP_START, OVERLAP_END)
    write_mortgage_ledger(output_root.parent / GENERATED_LEDGER_DIR / f"{year}-mortgage.beancount", year)
    validate(output_root, year)


def validate(output_root: Path, year: int) -> None:
    for provider, suffix in {"sparebank1": ".csv", "dnb": ".xlsx", "amex": ".qbo"}.items():
        expected = sorted([f"{year}-{month:02d}{suffix}" for month in range(1, 13)] + [f"{OVERLAP_START.isoformat()}_to_{OVERLAP_END.isoformat()}{suffix}"])
        files = sorted(path.name for path in (output_root / provider).iterdir())
        if files != expected:
            raise RuntimeError(f"unexpected {provider} files: {files}")
    for path in (output_root / "sparebank1").glob("*.csv"):
        with path.open(encoding="utf-8", newline="") as file:
            reader = csv.reader(file, delimiter=";")
            if next(reader) != ["Dato", "Beskrivelse", "Rentedato", "Inn", "Ut", "Til konto", "Fra konto", ""]:
                raise RuntimeError(f"unexpected CSV header in {path}")
            if not list(reader):
                raise RuntimeError(f"empty CSV export: {path}")
    for path in (output_root / "dnb").glob("*.xlsx"):
        ws = load_workbook(path, read_only=True).active
        if [cell.value for cell in next(ws.iter_rows(max_row=1))] != ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]:
            raise RuntimeError(f"unexpected XLSX header in {path}")
        if ws.max_row < 2:
            raise RuntimeError(f"empty XLSX export: {path}")
    for path in (output_root / "amex").glob("*.qbo"):
        root = ET.fromstring(path.read_text(encoding="utf-8").split("?>", maxsplit=2)[-1].encode())
        if not root.findall(".//STMTTRN"):
            raise RuntimeError(f"empty QBO export: {path}")
    mortgage_path = output_root.parent / GENERATED_LEDGER_DIR / f"{year}-mortgage.beancount"
    mortgage_text = mortgage_path.read_text(encoding="utf-8")
    if "Extra Mortgage Payment" not in mortgage_text or "Expenses:Interest:Mortgage" not in mortgage_text:
        raise RuntimeError(f"unexpected mortgage ledger content in {mortgage_path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, default=YEAR)
    parser.add_argument("--output-root", type=Path, default=Path("data"))
    args = parser.parse_args()
    emit(args.output_root, args.year)


if __name__ == "__main__":
    main()
