from __future__ import annotations

import csv
import os
import subprocess
import sys
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
GENERATED = ROOT / "generated"


def test_demo_data_has_full_year_and_single_overlap_export() -> None:
    expected = sorted([f"2025-{month:02d}" for month in range(1, 13)] + ["2025-02-15_to_2025-04-15"])

    assert sorted(path.stem for path in (DATA / "sparebank1").glob("*.csv")) == expected
    assert sorted(path.stem for path in (DATA / "dnb").glob("*.xlsx")) == expected
    assert sorted(path.stem for path in (DATA / "amex").glob("*.qbo")) == expected


def test_demo_data_provider_files_are_structurally_valid() -> None:
    with (DATA / "sparebank1" / "2025-01.csv").open(encoding="utf-8", newline="") as file:
        reader = csv.reader(file, delimiter=";")
        assert next(reader) == ["Dato", "Beskrivelse", "Rentedato", "Inn", "Ut", "Til konto", "Fra konto", ""]
        assert sum(1 for _ in reader) >= 12

    workbook = load_workbook(DATA / "dnb" / "2025-01.xlsx", read_only=True)
    worksheet = workbook.active
    assert [cell.value for cell in next(worksheet.iter_rows(max_row=1))] == ["Dato", "Beløpet gjelder", "Valuta", "Kurs", "Inn", "Ut"]
    assert worksheet.max_row >= 9

    qbo = (DATA / "amex" / "2025-01.qbo").read_text(encoding="utf-8")
    root = ET.fromstring(qbo.split("?>", maxsplit=2)[-1])
    assert len(root.findall(".//STMTTRN")) >= 6


def test_demo_data_includes_generated_mortgage_accounting() -> None:
    mortgage = (GENERATED / "2025-mortgage.beancount").read_text(encoding="utf-8")

    assert "Liabilities:Loan:Mortgage" in mortgage
    assert "Expenses:Interest:Mortgage" in mortgage
    assert "Extra Mortgage Payment - Vacation bonus" in mortgage
    assert "Extra Mortgage Payment - Year-end savings" in mortgage


def test_demo_ledger_declares_split_person() -> None:
    ledger = (ROOT / "main.beancount").read_text(encoding="utf-8")

    assert "open Assets:Receivable:Maria NOK" in ledger
    assert 'custom "split-person" "maria" Assets:Receivable:Maria' in ledger


def test_readme_documents_split_workflow() -> None:
    readme = (ROOT / "README.md").read_text(encoding="utf-8")

    assert "The annotation describes the other person's owed share" in readme
    assert 'split: "maria:50%"' in readme
    assert 'split: "maria:100%"' in readme
    assert (
        "uv run preserve-splits imports/2025.beancount "
        "imports/2025.fresh.beancount > imports/2025.preserved.beancount"
    ) in readme
    assert "mv imports/2025.preserved.beancount imports/2025.beancount" in readme
    assert (
        "uv run generate-splits --config main.beancount --year 2025 "
        "imports/2025.beancount > generated/2025-splits.beancount"
    ) in readme
    assert "Assets:Receivable:Maria          -446.15 NOK" in readme
    assert "exact amount splits" in readme
    assert "recurring/default\nsplit rules" in readme
    assert "refund semantics" in readme
    assert "automatic settlement matching" in readme
    assert "annotation helpers" in readme


def test_generated_demo_data_imports_with_configured_importers() -> None:
    result = subprocess.run(
        [sys.executable, "-m", "beancounters.importers", "extract", str(DATA)],
        cwd=ROOT,
        env={**os.environ, "PYTHONPATH": str(ROOT / "src")},
        text=True,
        capture_output=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr + result.stdout
    assert "Assets:Bank:SpareBank1:Checking" in result.stdout
    assert "Liabilities:CreditCard:DNB" in result.stdout
    assert "Liabilities:CreditCard:Amex" in result.stdout


def test_demo_query_script_reports_loan_insights(tmp_path: Path) -> None:
    output = tmp_path / "queries.md"
    result = subprocess.run(
        ["bash", "scripts/run-demo-queries.sh", "main.beancount", str(output)],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr + result.stdout
    report = output.read_text(encoding="utf-8")
    assert "## Loan Insights" in report
    assert "How Much Do I Save by Repaying Loan?" in report
    assert "Extra Mortgage Payment - Vacation bonus" in report
    assert "Loan Principal Paid by Month" in report
    assert "Liabilities:Loan:Mortgage" in report
